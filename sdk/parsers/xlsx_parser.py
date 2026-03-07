"""
Title: XLSX Parser
Abstract: Extracts metadata and sample data from Excel spreadsheets.
Why: Provides a structured summary of large workbooks, enabling LLMs to understand tabular data without processing millions of rows.
Dependencies: openpyxl, os, logging, sdk.parsers.base_parser, sdk.exceptions
"""

import os
import logging
from typing import Dict, Any, List, Set
import openpyxl
from sdk.parsers.base_parser import BaseParser
from sdk.exceptions import ParserError

logger = logging.getLogger(__name__)

class XlsxParser(BaseParser):
    """
    Parser for Excel files.
    Follows Pillar 1 (Strict Typing) and Pillar 2 (Interface-Driven).
    """

    def extract(self, file_path: str, **kwargs: Any) -> Dict[str, Any]:
        """
        Extracts content from an XLSX file.
        
        Args:
            file_path: Path to the .xlsx file.
            **kwargs: max_rows (default 5).
            
        Returns:
            Dict[str, Any]: {'text': str}
            
        Raises:
            ParserError: If extraction fails.
        """
        self._validate_file(file_path)
        max_rows = kwargs.get("max_rows", 5)

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet_names = workbook.sheetnames
            text = self._process_workbook(workbook, sheet_names, max_rows)
            return {"text": text}
        except Exception as e:
            logger.error(f"Failed to extract XLSX from {file_path}: {e}")
            raise ParserError(f"Error parsing XLSX: {e}") from e

    def _validate_file(self, file_path: str) -> None:
        """Validates file existence."""
        if not os.path.exists(file_path):
            raise ParserError(f"Excel file not found: {file_path}")

    def _process_workbook(self, workbook: openpyxl.Workbook, sheet_names: List[str], max_rows: int) -> str:
        """Summarizes the workbook content."""
        summary_parts = [f"Workbook contains {len(sheet_names)} sheets: {', '.join(sheet_names)}"]
        sampled_sheets = self._get_sampled_sheets(sheet_names)
        
        for sheet_name in sorted(list(sampled_sheets), key=lambda x: sheet_names.index(x)):
            sheet = workbook[sheet_name]
            summary_parts.append(self._summarize_sheet(sheet, sheet_name, max_rows))
            
        return "\n".join(summary_parts)

    def _get_sampled_sheets(self, sheet_names: List[str]) -> Set[str]:
        """Selects a representative sample of sheets."""
        total = len(sheet_names)
        sampled = {sheet_names[0], sheet_names[total // 2], sheet_names[-1]}
        if total > 1:
            sampled.add(sheet_names[1])
        if total > 2:
            sampled.add(sheet_names[total - 2])
        return sampled

    def _summarize_sheet(self, sheet: Any, sheet_name: str, max_rows: int) -> str:
        """Extracts summary and sample data from a single sheet."""
        rows = sheet.max_row if sheet.max_row else 0
        cols = sheet.max_column if sheet.max_column else 0
        lines = [f"Sheet '{sheet_name}': {rows} rows, {cols} columns."]
        
        sample_data = []
        for row in sheet.iter_rows(max_row=max_rows, values_only=True):
            row_str = " | ".join([str(cell) for cell in row if cell is not None])
            if row_str.strip():
                sample_data.append(row_str)
        
        if sample_data:
            lines.append("Sample rows:")
            lines.extend(sample_data)
        return "\n".join(lines)

if __name__ == "__main__":
    # Example usage
    logging.basicConfig(level=logging.INFO)
    parser = XlsxParser()
    print("XlsxParser initialized. Call extract(file_path) to use.")
